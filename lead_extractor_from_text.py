import re
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import spacy
except ImportError:
    print("Erro: Bibliotecas necessárias não instaladas.")
    print("Certifique-se de rodar: pip install openpyxl spacy")
    print("E também o modelo: python -m spacy download pt_core_news_sm")
    sys.exit()

BLACKLIST = {
    "Diretoria", "Suporte", "Erro", "Atenção", "Nota", "Boleto", "Sistema",
    "Tickets", "Interno", "Assunto", "Instalação", "O", "A", "Os", "As",
    "Olá", "Bom", "Dia", "Fim", "Arquivo", "Logs", "Financeiro", "Logística",
    "Recursos", "Humanos", "Infraestrutura", "Rede", "Equipe", "Usuário",
    "Cliente", "Motorista", "Tratar", "Por", "Favor", "Obrigado", "Obrigada",
    "Sr", "Sra", "Dr", "Dra", "Prof", "Profa", "Eng", "Enga", "Saudações",
    "Prezado", "Prezada", "Caro", "Cara", "Estimado", "Estimada"
}

def is_nome_valido(nome):
    nome = nome.strip()

    if len(nome) < 5 or len(nome) > 100:
        return False

    palavras = nome.split()

    if len(palavras) < 2 or len(palavras) > 5:
        return False

    for palavra in palavras:
        if len(palavra) < 2 or len(palavra) > 30:
            return False

    for palavra in palavras:
        if not palavra[0].isupper():
            return False
        for letra in palavra[1:]:
            if letra.isupper():
                return False

    if any(char.isdigit() for char in nome):
        return False

    if re.search(r'[!@#$%^&*()_+=\[\]{};:""\\|,.<>/?]', nome):
        return False

    palavras_invalidas = {"E", "De", "Da", "Do", "Das", "Dos", "A", "O", "As", "Os"}
    if any(palavra in palavras_invalidas for palavra in palavras):
        if len(palavras) <= 2:
            return False

    for palavra in palavras:
        if palavra in BLACKLIST:
            return False

    return True

def extrair_dados_blindados():
    nome_arquivo_txt = "texto_leitura.txt"
    nome_arquivo_excel = "dados_extraidos.xlsx"

    if not os.path.exists(nome_arquivo_txt):
        print(f"Erro: O arquivo '{nome_arquivo_txt}' não foi encontrado.")
        return

    with open(nome_arquivo_txt, "r", encoding="utf-8") as arquivo:
        texto_exterior = arquivo.read()

    try:
        nlp = spacy.load("pt_core_news_sm")
    except Exception:
        print("Erro ao carregar o modelo do SpaCy. Rode no terminal: python -m spacy download pt_core_news_sm")
        return

    doc = nlp(texto_exterior)

    nomes_detectados_pela_ia = []
    for entidade in doc.ents:
        if entidade.label_ == "PER":
            nome_limpo = entidade.text.strip()
            nome_limpo = re.sub(r'\s+', ' ', nome_limpo)
            nomes_detectados_pela_ia.append(nome_limpo)

    nomes_finais = []
    nomes_rejeitados = []

    for nome in nomes_detectados_pela_ia:
        if is_nome_valido(nome):
            nomes_finais.append(nome)
        else:
            nomes_rejeitados.append(nome)

    regex_email = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    regex_telefone = r'\(?\d{2,3}\)?[\s-]?\d{4,5}[\s-]?\d{4}|\d{7,11}'

    emails = sorted(list(set(re.findall(regex_email, texto_exterior))))
    telefones = sorted(list(set(re.findall(regex_telefone, texto_exterior))))
    nomes = sorted(list(set(nomes_finais)))

    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws = wb.create_sheet(title="Dados Extraídos")

    fonte_cabecalho = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center")

    cor_nome = "4F81BD"
    cor_telefone = "9BBB59"
    cor_email = "F79646"

    cabecalhos = ["NOMES", "TELEFONES", "E-MAILS"]

    for col, cabecalho in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=1, column=col, value=cabecalho)
        celula.font = fonte_cabecalho
        celula.alignment = alinhamento_centro

        if col == 1:
            celula.fill = PatternFill(start_color=cor_nome, end_color=cor_nome, fill_type="solid")
        elif col == 2:
            celula.fill = PatternFill(start_color=cor_telefone, end_color=cor_telefone, fill_type="solid")
        elif col == 3:
            celula.fill = PatternFill(start_color=cor_email, end_color=cor_email, fill_type="solid")

    max_linhas = max(len(nomes), len(telefones), len(emails))

    for linha in range(max_linhas):
        if linha < len(nomes):
            ws.cell(row=linha + 2, column=1, value=nomes[linha]).alignment = alinhamento_esquerda
        else:
            ws.cell(row=linha + 2, column=1, value="").alignment = alinhamento_esquerda

        if linha < len(telefones):
            ws.cell(row=linha + 2, column=2, value=telefones[linha]).alignment = alinhamento_esquerda
        else:
            ws.cell(row=linha + 2, column=2, value="").alignment = alinhamento_esquerda

        if linha < len(emails):
            ws.cell(row=linha + 2, column=3, value=emails[linha]).alignment = alinhamento_esquerda
        else:
            ws.cell(row=linha + 2, column=3, value="").alignment = alinhamento_esquerda

    max_len_nomes = max([len(str(nome)) for nome in nomes]) if nomes else 0
    ws.column_dimensions['A'].width = max(max_len_nomes + 5, 25)

    max_len_telefones = max([len(str(tel)) for tel in telefones]) if telefones else 0
    ws.column_dimensions['B'].width = max(max_len_telefones + 5, 20)

    max_len_emails = max([len(str(email)) for email in emails]) if emails else 0
    ws.column_dimensions['C'].width = max(max_len_emails + 5, 30)

    wb.save(nome_arquivo_excel)

    print(f"\n[SUCESSO] Processamento com IA e validação rigorosa concluído!")
    print(f"Resultados limpos: {len(nomes)} nomes, {len(telefones)} telefones, {len(emails)} e-mails.")

    if nomes_rejeitados:
        print(f"\n[INFO] Nomes rejeitados pela validação rigorosa: {len(nomes_rejeitados)}")
        print("Exemplos de nomes rejeitados (primeiros 5):")
        for nome in nomes_rejeitados[:5]:
            print(f"  - {nome}")

    print(f"\nArquivo salvo como: {nome_arquivo_excel}")
    print("Os dados foram organizados em uma única aba com colunas ajustadas.")

if __name__ == "__main__":
    extrair_dados_blindados()
